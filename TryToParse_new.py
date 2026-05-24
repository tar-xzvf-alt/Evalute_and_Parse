import sys

import customtkinter as ctk
import parsar
import threading
from tkinter import *
import datetime as DT
import time
import os
from PIL import Image
from customtkinter import CTkImage
from tkinter import filedialog, messagebox
import slovari_script_new
import slovari_parse_itogi
from openpyxl import Workbook
import pandas as pd

df_left = []

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def show_graph(graph_path):
    global graph_img_label
    image = Image.open(resource_path(graph_path))
    img = CTkImage(light_image=image, size=(600, 400))
    graph_img_label.configure(image=img, text="")
    graph_img_label.image = img

def analyse_file():
    analyse_button.configure(state=DISABLED)
    folder_entry.configure(state="readonly")

    file_path = file_entry.get()
    folder_path = folder_entry.get()

    if not file_path.endswith(".xlsx"):
        messagebox.showerror("Ошибка", "Пожалуйста, выберите файл формата .xlsx")
        return

    file_name = os.path.basename(file_path)
    #folder_path = os.path.dirname(file_path)

    print(f"Обрабатываем файл: {file_name}")

    use_first = use_dict1.get()
    use_second = use_dict2.get()

    if not use_first and not use_second:
        messagebox.showerror("Ошибка", "Выберите хотя бы один словарь")
        return

    progressbar_analyse.start()
    status_label_analyse.configure(text=f"Сбор информации...")

    try:
        df = pd.read_excel(file_path)
        if df.empty:
            messagebox.showwarning("Пустой файл", "Файл не содержит данных")
            return

        df['text'] = df[df.columns[0]].astype(str)
        df = df[df['text'].str.strip().astype(bool)]

        results = []
        titles = []

        if use_first:
            df1 = df.copy()
            df1["sentiment"] = df1["text"].apply(lambda t: slovari_script_new.determine_sentiment_score(t, dict_id=1))
            results.append(df1)
            titles.append("Словарь 1")

            output_name_1 = os.path.splitext(file_name)[0]
            output_path_1 = os.path.join(folder_path, f"{output_name_1}_1_разметка.xlsx")
            df1.to_excel(output_path_1, index=False)

            total_1 = len(df)
            positive_1 = (df1["sentiment"] == 1.0).sum()
            negative_1 = (df1["sentiment"] == 0.0).sum()

            wb = Workbook()
            ws = wb.active
            stats = [
                ("Файл", output_name_1),
                ("Всего текстов", total_1),
                ("Негативных", negative_1),
                ("Позитивных", positive_1),
                ("% негативных", negative_1 / total_1 * 100),
                ("% позитивных", positive_1 / total_1 * 100)
            ]
            for row, (label, value) in enumerate(stats, start=1):
                ws[f"A{row}"] = label
                ws[f"B{row}"] = value

            stats_path = os.path.join(folder_path, f"{output_name_1}_1_итоги.xlsx")
            wb.save(stats_path)


        if use_second:
            df2 = df.copy()
            df2["sentiment"] = df2["text"].apply(lambda t: slovari_script_new.determine_sentiment_score(t, dict_id=2))
            results.append(df2)
            titles.append("Словарь 2")

            output_name_2 = os.path.splitext(file_name)[0]
            output_path_2 = os.path.join(folder_path, f"{output_name_2}_2_разметка.xlsx")
            df2.to_excel(output_path_2, index=False)

            total_2 = len(df)
            positive_2 = (df2["sentiment"] == 1.0).sum()
            negative_2 = (df2["sentiment"] == 0.0).sum()

            wb = Workbook()
            ws = wb.active
            stats = [
                ("Файл", output_name_2),
                ("Всего текстов", total_2),
                ("Негативных", negative_2),
                ("Позитивных", positive_2),
                ("% негативных", negative_2 / total_2 * 100),
                ("% позитивных", positive_2 / total_2 * 100)
            ]
            for row, (label, value) in enumerate(stats, start=1):
                ws[f"A{row}"] = label
                ws[f"B{row}"] = value

            stats_path = os.path.join(folder_path, f"{output_name_2}_2_итоги.xlsx")
            wb.save(stats_path)

        progressbar_analyse.stop()
        analyse_button.configure(state=NORMAL)

        status_label_analyse.configure(text=f"Ожидание запуска...")

        # Визуализация одной или двух диаграмм
        plot_path = slovari_script_new.visualize_multiple(results, titles, file_path)
        show_graph(plot_path)
        folder_entry.configure(state="normal")


    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")


def get_and_parse_theme(alt_themes, all_keywords, year,date1, date2, filename):

    if (all_keywords != '') and (date1 != '') and (date2 != '') and (alt_themes != '') and (entry_group_count.get() != '') and (
            entry_min_words.get() != '') and not (checkbox_posts == 'off' and checkbox_comments == 'off')and(
            entry_token.get() != ''):
        dt1 = DT.datetime.strptime(date1, '%d.%m.%Y').timestamp()
        dt2 = DT.datetime.strptime(date2, '%d.%m.%Y').timestamp()

        status_label.configure(text=f"Ожидание...")
        status_label.pack()
        if date2 < date1:
            status_label.configure(text=f"Введите корректные даты!")
            status_label.pack()

        Comments_By_Theme = []
        Posts_By_Theme = []


        btn_start.configure(state=DISABLED)
        entry_keywords_search.configure(state=DISABLED)
        entry_from.configure(state=DISABLED)
        entry_to.configure(state=DISABLED)
        entry_keywords.configure(state=DISABLED)
        entry_group_count.configure(state=DISABLED)
        entry_min_words.configure(state=DISABLED)
        checkbox_posts.configure(state=DISABLED)
        checkbox_comments.configure(state=DISABLED)

        all_words_for_search = alt_themes.split(',')

        theme_part_of_filename = "_".join(all_words_for_search)


        status_label.configure(text="Данные считаны")

        progressbar.start()

        PostsCount = 0
        CommentsCount = 0

        if checkbox_posts.get() == 0 and checkbox_comments.get() == 1:
            for word_in_search_keywords in all_words_for_search:
                temp = GetAndParseComments(word_in_search_keywords)
                Comments_By_Theme.extend(temp)
                CommentsCount += len(temp)

            status_label.configure(text=f"Печать комментариев")
            time.sleep(2)

            if year == '':
                parsar.file_writer(Comments_By_Theme, f"{theme_part_of_filename}_{date1}_{date2}_комментарии", dt1, dt2,filename)
                show_result_only_comments(f"{theme_part_of_filename}_{date1}_{date2}_комментарии", len(Comments_By_Theme), status_label)

            else:
                parsar.file_writer(Comments_By_Theme, f"{theme_part_of_filename}_{year}_комментарии", dt1, dt2,filename)
                show_result_only_comments(f"{theme_part_of_filename}_{year}_комментарии", len(Comments_By_Theme), status_label)

            progressbar.stop()
        else:
            if checkbox_posts.get() == 1 and checkbox_comments.get() == 0:
                for word_in_search_keywords in all_words_for_search:
                    temp = GetAndParsePosts(word_in_search_keywords)
                    Posts_By_Theme.extend(temp)
                    PostsCount += len(temp)

                status_label.configure(text=f"Печать постов")
                time.sleep(2)
                if year == '':
                    parsar.file_writer(Posts_By_Theme, f"{theme_part_of_filename}_{date1}_{date2}_посты", dt1,
                                       dt2, filename)
                else:
                    parsar.file_writer(Posts_By_Theme, f"{theme_part_of_filename}_{year}_посты", dt1, dt2,
                                       filename)
                status_label.configure(text=f"Ожидание запуска")
                progressbar.stop()
                ShowResultOnlyPosts(alt_themes, PostsCount, status_label)
            else:
                if checkbox_posts.get() == 1 and checkbox_comments.get() == 0:
                    for word_in_search_keywords in all_words_for_search:
                        temp_posts = GetAndParsePosts(word_in_search_keywords)
                        temp_comments = GetAndParseComments(word_in_search_keywords)

                        Posts_By_Theme.extend(temp_posts)
                        Comments_By_Theme.extend(temp_comments)

                        PostsCount+= len(temp_posts)
                        CommentsCount += len(temp_comments)

                    status_label.configure(text=f"Печать постов")
                    time.sleep(2)
                    if year == '':
                        parsar.file_writer(Posts_By_Theme, f"{theme_part_of_filename}_{date1}_{date2}_посты",
                                           dt1, dt2, filename)
                        ShowResult(f"{theme_part_of_filename}_{date1}_{date2}_посты", PostsCount, CommentsCount, status_label)

                    else:
                        parsar.file_writer(Posts_By_Theme, f"{theme_part_of_filename}_{year}_посты", dt1, dt2,
                                           filename)
                        ShowResult(f"{theme_part_of_filename}_{year}_посты", PostsCount, CommentsCount, status_label)



                    status_label.configure(text=f"Печать комментариев")
                    time.sleep(2)
                    if year == '':
                        parsar.file_writer(Comments_By_Theme, f"{theme_part_of_filename}_{date1}_{date2}_комментарии",
                                           dt1, dt2, filename)
                        ShowResult(f"{theme_part_of_filename}_{date1}_{date2}_комментарии", PostsCount, CommentsCount, status_label)

                    else:
                        parsar.file_writer(Comments_By_Theme, f"{theme_part_of_filename}_{year}_комментарии", dt1, dt2,
                                           filename)
                        ShowResult(f"{theme_part_of_filename}_{year}_комментарии", PostsCount, CommentsCount, status_label)

                    progressbar.stop()
                    ShowResult(alt_themes, PostsCount, CommentsCount, status_label)

        btn_start.configure(state=NORMAL)
        entry_keywords_search.configure(state=NORMAL)
        entry_from.configure(state=NORMAL)
        entry_to.configure(state=NORMAL)
        entry_keywords.configure(state=NORMAL)
        entry_group_count.configure(state=NORMAL)
        entry_min_words.configure(state=NORMAL)
        checkbox_posts.configure(state=NORMAL)
        checkbox_comments.configure(state=NORMAL)
    else:
        status_label.configure(text=f"Данных недостаточно")
        status_label.pack(pady=5)


def show_result_only_comments(theme, comments_count, status_label):
    status_label.configure(text=f"Файл \n {theme}.xlsx создан\nОтобрано {comments_count} комментариев")
    status_label.pack(pady=5)


def ShowResult(theme, posts_count, comments_count, status_label):
    status_label.configure(tab_parse,
                     text=f"Файл \n {theme}.xlsx создан\nОтобрано\n{posts_count} постов\n{comments_count} комментариев")
    status_label.pack(pady=5)


def ShowResultOnlyPosts(theme, posts_count, status_label):
    status_label.configure(text=f"Файл \n {theme}.xlsx создан\nОтобрано {posts_count} постов")
    status_label.pack(pady=5)


def GetAndParsePosts(one_theme):
    filename = entry_folder_parse.get()
    date1 = entry_from.get()
    date2 = entry_to.get()
    city = one_theme
    count = entry_group_count.get()
    minimal = entry_min_words.get()
    words = entry_keywords.get()

    if (city != '') and (date1 != '') and (date2 != '') and (filename != '') and (count != '') and (minimal != ''):
        dt1 = DT.datetime.strptime(date1, '%d.%m.%Y')
        dt2 = DT.datetime.strptime(date2, '%d.%m.%Y')

        status_label.configure(text=f"Парсинг в сообществах близких к теме {one_theme}...")
        PostsByCity = parsar.MainParsePosts(city, dt1.timestamp(), dt2.timestamp(), count, filename, minimal, words,entry_token.get())
        status_label.configure(text=f"Отобрано {len(PostsByCity)} постов")
        time.sleep(2)
        return PostsByCity

    else:
        status_label.configure(text=f"Данных недостаточно")
        status_label.pack()
        return []


def GetAndParseComments(one_theme):

    filename = entry_folder_parse.get()
    date1 = entry_from.get()
    date2 = entry_to.get()
    theme = one_theme
    count = entry_group_count.get()
    minimal = entry_min_words.get()
    words = entry_keywords.get()

    if (theme != '') and (date1 != '') and (date2 != '') and (filename != '') and (count != '') and (minimal != ''):
        dt1 = DT.datetime.strptime(date1, '%d.%m.%Y')
        dt2 = DT.datetime.strptime(date2, '%d.%m.%Y')

        status_label.configure(text=f"Сбор комментов по теме {theme}...")
        CommentsByCity = parsar.MainParseComments(theme, dt1.timestamp(), dt2.timestamp(), count, filename, minimal, words,entry_token.get())
        status_label.configure(text=f"Собрано {len(CommentsByCity)} комментов")
        time.sleep(2)

        return CommentsByCity

    else:
        status_label.configure(text=f"Данных недостаточно")
        status_label.pack()
        return []


def choose_file():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
    )
    if file_path:
        file_entry.configure(state="normal")
        file_entry.delete(0, "end")
        file_entry.insert(0, file_path)
        file_entry.configure(state="readonly")


def choose_dir():
    folder_path = filedialog.askdirectory()
    folder_entry.delete(0, "end")
    folder_entry.insert(0, folder_path)


def choose_dir_parse():
    folder_path = filedialog.askdirectory()
    entry_folder_parse.insert(0, folder_path)
    entry_folder_parse.configure(state="readonly")


def choose_dir_spec():
    folder_path = filedialog.askdirectory()
    folder_entry_spec.insert(0, folder_path)
    folder_entry_spec.configure(state="readonly")


def analyse_global():
    base_folder = folder_entry_spec.get()
    if not os.path.isdir(base_folder):
        messagebox.showerror("Ошибка", "Пожалуйста, выберите существующую папку")
        return

    progressbar_analyse_spec.start()
    status_label_analyse_spec.configure(text="Сбор итогов...")

    slovari_parse_itogi.get_all_info_about_folders(base_folder)

    status_label_analyse_spec.configure(text="Сбор итогов окончен!")
    progressbar_analyse_spec.stop()






# Настройки темы
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Главное окно
root = ctk.CTk()

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

width,height = 700,800

# Вычисляем позицию для центрирования
x = (screen_width // 2) - (width // 2)
y = (screen_height // 2) - (height // 2)


root.geometry(f"{width}x{height}+{x}+{y}")

root.title("ParseVK")

# Вкладки
tabview = ctk.CTkTabview(master=root, width=680, height=680)
tabview.pack(padx=10, pady=10)
tabview.add("Парсинг")
tabview.add("API Токен")
tabview.add("Анализ среза")
tabview.add("Мониторинг срезов")


tab_parse = tabview.tab("Парсинг")
tab_token = tabview.tab("API Токен")
tab_stats = tabview.tab("Анализ среза")
tab_stats_spec = tabview.tab("Мониторинг срезов")

# ============ Вкладка "Парсинг" ============

# Период
ctk.CTkLabel(tab_parse, text="Укажите период", font=("Arial", 14)).pack(pady=(15, 5))
frame_dates = ctk.CTkFrame(tab_parse, fg_color="transparent")
frame_dates.pack(pady=5)

entry_from = ctk.CTkEntry(frame_dates, placeholder_text="От (01.01.2022)", width=200)
entry_to = ctk.CTkEntry(frame_dates, placeholder_text="До (01.01.2023)", width=200)

# Метка и поле для года
label_year = ctk.CTkLabel(frame_dates, text="Указать год:")
entry_year = ctk.CTkEntry(frame_dates, placeholder_text="2022", width=80)

# Размещаем элементы в сетке
entry_from.grid(row=0, column=1, padx=(0, 20), pady=5, sticky="w")
entry_to.grid(row=0, column=3, padx=(0, 20), pady=5, sticky="w")

label_year.grid(row=0, column=4, padx=(20, 5), pady=5, sticky="e")
entry_year.grid(row=0, column=5, padx=(0, 5), pady=5, sticky="w")

def update_dates_from_year():
    year = entry_year.get()
    if year.isdigit() and len(year) == 4:
        entry_from.delete(0, "end")
        entry_from.insert(0, f"01.01.{year}")
        entry_to.delete(0, "end")
        entry_to.insert(0, f"01.01.{int(year)+1}")

# Привязываем обработчик события, когда поле года теряет фокус
entry_year.bind("<FocusOut>", lambda e: update_dates_from_year())

# Выборка
ctk.CTkLabel(tab_parse, text="Выборка", font=("Arial", 14)).pack(pady=(10, 5))

entry_keywords_search = ctk.CTkEntry(tab_parse, placeholder_text="Ключевые слова для поиска сообществ", width=500)
entry_keywords_search.pack(pady=5)

entry_group_count = ctk.CTkEntry(tab_parse, placeholder_text="Количество сообществ", width=500)
entry_group_count.pack(pady=5)

entry_min_words = ctk.CTkEntry(tab_parse, placeholder_text="Мин. количество слов", width=500)
entry_min_words.pack(pady=5)

entry_keywords = ctk.CTkEntry(tab_parse, placeholder_text="Ключевые слова", width=500)
entry_keywords.pack(pady=5)

# Запись в файл
ctk.CTkLabel(tab_parse, text="Запись в файл", font=("Arial", 14)).pack(pady=(10, 5))
# Рамка для выбора папки
frame_folder = ctk.CTkFrame(tab_parse)
frame_folder.pack(pady=10, padx=10, fill="x")

entry_folder_parse = ctk.CTkEntry(frame_folder, width=400, placeholder_text="Выберите папку для сохранения итогов")
entry_folder_parse.pack(side="left", padx=(10, 5), pady=10)

folder_button = ctk.CTkButton(frame_folder, text="Обзор", command=choose_dir_parse)
folder_button.pack(side="left", padx=5, pady=10)

# Чекбоксы
frame_checks = ctk.CTkFrame(tab_parse, fg_color="transparent")
frame_checks.pack(pady=5)
var_posts = ctk.BooleanVar(value=False)
var_comments = ctk.BooleanVar(value=False)
checkbox_posts = ctk.CTkCheckBox(frame_checks, text="Посты", variable=var_posts)
checkbox_comments = ctk.CTkCheckBox(frame_checks, text="Комментарии", variable=var_comments)
checkbox_posts.pack(side="left", padx=10, pady=5)
checkbox_comments.pack(side="left", padx=10, pady=5)

# Прогресс
progressbar = ctk.CTkProgressBar(tab_parse, width=400)
progressbar.set(0.5)
progressbar.pack(pady=10)

# Статус
status_label = ctk.CTkLabel(tab_parse, text="Ожидание запуска...", text_color="gray")
status_label.pack()

# Кнопка запуска
btn_start = ctk.CTkButton(tab_parse, text="Запуск",
                    # bg_color='orange',
                    command=lambda: threading.Thread(target=get_and_parse_theme,
                                                     args=(entry_keywords_search.get(), entry_keywords.get(), entry_year.get(),entry_from.get(), entry_to.get(), entry_folder_parse.get()),
                                                     daemon=True).start(),
                    font=('Arial', 16, 'bold'),
                    width=70,
                    height=30
                    )
btn_start.pack(pady=10)


# ============ Вкладка "Статистика" ============
label_stats = ctk.CTkLabel(tab_stats, text="Анализ собранной информации", font=("Arial", 14))
label_stats.pack(pady=20)

# Выбор словарей
dict_select_frame = ctk.CTkFrame(tab_stats)
dict_select_frame.pack(pady=5)

use_dict1 = ctk.BooleanVar(value=True)
use_dict2 = ctk.BooleanVar(value=False)

checkbox_dict1 = ctk.CTkCheckBox(dict_select_frame, text="Словарь 1", variable=use_dict1)
checkbox_dict2 = ctk.CTkCheckBox(dict_select_frame, text="Словарь 2", variable=use_dict2)

checkbox_dict1.pack(side="left", padx=10)
checkbox_dict2.pack(side="left", padx=10)

# Рамка для выбора файла
frame_file = ctk.CTkFrame(tab_stats)
frame_file.pack(pady=10, padx=10, fill="x")

file_entry = ctk.CTkEntry(frame_file, width=400, placeholder_text="Выберите файл для анализа")
file_entry.pack(side="left", padx=(10, 5), pady=5)

browse_button = ctk.CTkButton(frame_file, text="Обзор", command=choose_file)
browse_button.pack(side="left", padx=5, pady=5)

# Рамка для выбора папки
frame_folder = ctk.CTkFrame(tab_stats)
frame_folder.pack(pady=10, padx=10, fill="x")

folder_entry = ctk.CTkEntry(frame_folder, width=400, placeholder_text="Выберите папку для сохранения итогов")
folder_entry.pack(side="left", padx=(10, 5), pady=10)

folder_button = ctk.CTkButton(frame_folder, text="Обзор", command=choose_dir)
folder_button.pack(side="left", padx=5, pady=10)

# Прогресс
progressbar_analyse = ctk.CTkProgressBar(tab_stats, width=400)
progressbar_analyse.set(0.5)
progressbar_analyse.pack(pady=5)

# Статус
status_label_analyse = ctk.CTkLabel(tab_stats, text="Ожидание запуска...", text_color="gray")
status_label_analyse.pack()

# Кнопка для запуска анализа
analyse_button = ctk.CTkButton(tab_stats, text="Запустить анализ", command=lambda: threading.Thread(target=analyse_file, daemon=True).start())
analyse_button.pack(pady=10)

# Место под график
graph_img_label = ctk.CTkLabel(tab_stats, text="[ Здесь появится график ]", font=("Arial", 16), text_color="gray")
graph_img_label.pack(pady=20)


# ============ Вкладка "Сводный анализ" ============
label_stats = ctk.CTkLabel(tab_stats_spec, text="Сводный анализ информации", font=("Arial", 14))
label_stats.pack(pady=20)

# Рамка для выбора файла
frame_file_spec = ctk.CTkFrame(tab_stats_spec)
frame_file_spec.pack(pady=10, padx=10, fill="x")

folder_entry_spec = ctk.CTkEntry(frame_file_spec, width=400, placeholder_text="Выберите каталог для сводного анализа по файлам")
folder_entry_spec.pack(side="left", padx=(10, 5), pady=10)

browse_button_spec = ctk.CTkButton(frame_file_spec, text="Обзор", command=choose_dir_spec)
browse_button_spec.pack(side="left", padx=5, pady=10)

# Прогресс
progressbar_analyse_spec = ctk.CTkProgressBar(tab_stats_spec, width=400)
progressbar_analyse_spec.set(0.5)
progressbar_analyse_spec.pack(pady=5)

# Статус
status_label_analyse_spec = ctk.CTkLabel(tab_stats_spec, text="Ожидание запуска...", text_color="gray")
status_label_analyse_spec.pack()

# Кнопка для запуска анализа
analyse_button_spec = ctk.CTkButton(tab_stats_spec, text="Создать сводные файлы", command=lambda: threading.Thread(target=analyse_global, daemon=True).start())
analyse_button_spec.pack(pady=10)

# ============ Вкладка "API Токен" ============

# Функция для загрузки токена при старте

def load_token():
    token_file = resource_path('vk_token.txt')
    if os.path.exists(token_file):
        try:
            with open(token_file, 'r') as f:
                token = f.read().strip()
                if token and parsar.check_vk_token(token):
                    entry_token.insert(0, token)
                    return True
        except Exception as e:
            print(f"Ошибка при загрузке токена: {e}")
    return False

# Показ справки

def show_help():
    messagebox.showinfo(
        "Справка",
        "Для получения токена вам нужно нажать кнопку 'Получить токен' и следовать следующей инструкции:\n"
        "1. Нажмите кнопку 'Получить токен' для перехода к генератору токенов\n"
        "2. Авторизируйтесь в свою учетную запись ВКонтакте \n"
        "3. Скопируйте полученный токен из адресной строки символы после 'access_token=' до '&expires_in'\n"
        "4. Вставьте токен в поле ввода\n"
        "5. Проверьте токен кнопкой 'Проверить токен'\n"
        "6. Сохраните токен кнопкой 'Сохранить токен'"
    )

# Функция для сохранения токена
def save_token():
    token = entry_token.get().strip()
    if not token:
        messagebox.showerror("Ошибка", "Поле токена не может быть пустым!")
        return
    try:
        with open(resource_path('vk_token.txt'), 'w') as f:
            f.write(token)
        messagebox.showinfo("Успех", f"Токен успешно сохранен! {resource_path('vk_token.txt')}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить токен: {str(e)}")

# Функция для проверки токена
def validate_token():
    token = entry_token.get().strip()

    if not token:
        messagebox.showerror("Ошибка", "Поле токена не может быть пустым!")
        return

    if not token.startswith('vk1.a.') or len(token) < 50:
        messagebox.showerror("Ошибка", "Токен имеет неверный формат!")
        return

    if parsar.check_vk_token(token):
        messagebox.showinfo("Успех", "Токен действителен и работает!")
    else:
        messagebox.showerror("Ошибка", "Токен недействителен или не имеет доступа!")


# Получение токена
def open_link():
    import webbrowser                                         
    webbrowser.open("https://oauth.vk.com/authorize?client_id=54447901&display=page&redirect_uri=https://oauth.vk.com/blank.html&response_type=token&v=5.131")


label_token = ctk.CTkLabel(tab_token, text="Токен VK API", font=("Arial", 14))
label_token.pack(pady=5)
entry_token = ctk.CTkEntry(tab_token, width=500, placeholder_text="Введите ваш токен")
entry_token.pack(pady=5)

# Загружаем токен при инициализации
if load_token():
    messagebox.showinfo("Результат загрузки токена", "Валидный токен для работы уже загружен из сохранения")
else:
    messagebox.showinfo("Результат загрузки токена", "Токен для работы не был загружен из сохранения или не валиден.\n"
                                                     "Для получения, проверки и сохранения токена перейдите на соответствующую вкладку")

# Фрейм для кнопок сохранения и проверки
button_top_frame = ctk.CTkFrame(tab_token)
button_top_frame.pack(pady=10)

# Кнопка проверки
validate_button = ctk.CTkButton(
    button_top_frame,
    text="Проверить токен",
    command=validate_token,
    #fg_color="#2aa9ff"
)
validate_button.pack(side="left", padx=5)

# Кнопка сохранения
save_button = ctk.CTkButton(
    button_top_frame,
    text="Сохранить токен",
    command=save_token
)
save_button.pack(side="left", padx=5)


# Фрейм для нижних кнопок
button_bottom_frame = ctk.CTkFrame(tab_token)
button_bottom_frame.pack(padx=10, pady=10)


link_button = ctk.CTkButton(
    button_bottom_frame,
    text="Получить токен",
    command=open_link,
    fg_color="transparent",
    border_width=1,
    text_color=("gray10", "gray90")
)
link_button.pack(side="left", padx=5)


help_button = ctk.CTkButton(
    button_bottom_frame,
    text="Помощь",
    width=30,
    command=show_help,
    fg_color="transparent",
    border_width=1,
    text_color=("gray10", "gray90")
)
help_button.pack(side="right", padx=5)

# Запуск
#root.iconbitmap(resource_path('favicon.ico'))

root.mainloop()
