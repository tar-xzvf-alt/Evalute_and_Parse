# coding: utf-8

import pandas as pd
import os
import sys
import logging
from openpyxl import Workbook
import pymorphy3
import nltk
from nltk.corpus import stopwords
from datetime import datetime

# Скачиваем стоп-слова
nltk.download('stopwords')

# ===============================
# Универсальный путь к ресурсам
# ===============================
def resource_path(relative_path):
    """ Получает абсолютный путь к ресурсу (работает и из .py, и из .exe) """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# ===============================
# Настройка логов
# ===============================
def setup_logging(log_dir):
    """Настройка логирования"""
    log_file = os.path.join(log_dir, f'batch_sentiment_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    logging.basicConfig(
        filename=log_file,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return log_file

# Инициализация морфологического анализатора и стоп-слов
morph = pymorphy3.MorphAnalyzer()
russian_stopwords = set(stopwords.words("russian"))

# ===============================
# Загрузка словарей тональностей
# ===============================
def load_sentiment_dict(dict_id):
    """
    Загружает словарь тональностей
    dict_id: 1 - kartaslovsent, 2 - rusentilex
    """
    try:
        if dict_id == 1:
            csv_name = "data/kartaslovsent.csv"
            csv_path = resource_path(csv_name)
            sentiment_data = pd.read_csv(csv_path, usecols=['term', 'value'], sep=";")
            return dict(zip(sentiment_data["term"], sentiment_data["value"]))
        else:
            csv_name = "data/rusentilex.csv"
            csv_path = resource_path(csv_name)
            sentiment_data = pd.read_csv(csv_path, usecols=['term', 'value'], sep=",")
            return dict(zip(sentiment_data["term"], sentiment_data["value"]))
    except Exception as e:
        logging.error(f"Ошибка загрузки словаря {dict_id}: {e}")
        raise

# ===============================
# Определение тональности текста
# ===============================
def determine_sentiment_score(text, sentiment_dict):
    """Определяет тональность текста на основе словаря"""
    if not isinstance(text, str) or not text.strip():
        return 0.0

    words = text.split()
    score = 0
    count = 0

    for word in words:
        try:
            normal_form = morph.parse(word.lower())[0].normal_form
            if normal_form in sentiment_dict and normal_form not in russian_stopwords:
                score += sentiment_dict[normal_form]
                count += 1
        except:
            continue

    if count == 0:
        return 0.0

    average_score = score / count
    return 1.0 if average_score >= 0.2 else 0.0

# ===============================
# Обработка одного файла
# ===============================
def process_single_file(file_path, output_dir, dict_id, dict_name):
    """
    Обрабатывает один Excel файл с текстами
    
    Parameters:
    file_path: путь к исходному файлу
    output_dir: директория для сохранения результатов
    dict_id: идентификатор словаря (1 или 2)
    dict_name: название словаря для формирования имени файла
    
    Returns:
    tuple: (total_count, positive_count, negative_count)
    """
    try:
        # Читаем файл
        df = pd.read_excel(file_path)
        
        if df.empty:
            logging.warning(f"Файл пуст: {file_path}")
            return None
        
        # Проверяем наличие колонки text
        if 'text' not in df.columns:
            logging.error(f"В файле {file_path} нет колонки 'text'")
            return None
        
        # Очищаем данные
        df = df.dropna(subset=['text'])
        df['text'] = df['text'].astype(str)
        df = df[df['text'].str.strip().astype(bool)]
        
        # Загружаем словарь
        sentiment_dict = load_sentiment_dict(dict_id)
        
        # Размечаем тексты
        df["sentiment"] = df["text"].apply(lambda t: determine_sentiment_score(t, sentiment_dict))
        
        # Сохраняем размеченный файл
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_file = os.path.join(output_dir, f"{base_name}_разметка_{dict_name}.xlsx")
        df.to_excel(output_file, index=False)
        
        # Собираем статистику
        total = len(df)
        positive = (df["sentiment"] == 1.0).sum()
        negative = (df["sentiment"] == 0.0).sum()
        
        # Сохраняем статистику
        wb = Workbook()
        ws = wb.active
        stats = [
            ("Файл", base_name),
            ("Словарь", dict_name),
            ("Всего текстов", total),
            ("Негативных", negative),
            ("Позитивных", positive),
            ("% негативных", negative / total * 100 if total > 0 else 0),
            ("% позитивных", positive / total * 100 if total > 0 else 0)
        ]
        
        for row, (label, value) in enumerate(stats, start=1):
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
        
        stats_file = os.path.join(output_dir, f"{base_name}_итоги_{dict_name}.xlsx")
        wb.save(stats_file)
        
        logging.info(f"Обработан файл: {file_path} (текстов: {total}, позитивных: {positive}, негативных: {negative})")
        
        return (total, positive, negative)
        
    except Exception as e:
        logging.error(f"Ошибка обработки файла {file_path}: {e}")
        return None

# ===============================
# Обработка всех файлов в иерархии
# ===============================
def process_all_files_in_hierarchy(base_folder, output_base_folder=None):
    """
    Рекурсивно обрабатывает все Excel файлы в иерархии папок
    
    Parameters:
    base_folder: корневая папка с иерархией
    output_base_folder: базовая папка для сохранения результатов
    """
    # Создаем папку для логов
    log_dir = base_folder if output_base_folder is None else output_base_folder
    log_file = setup_logging(log_dir)
    print(f"Логи сохраняются в: {log_file}")
    
    # Определяем базовую выходную папку
    if output_base_folder is None:
        output_base_folder = os.path.join(base_folder, "размеченные_данные")
    
    # Создаем базовую выходную папку
    os.makedirs(output_base_folder, exist_ok=True)
    
    # Словари для сбора статистики по папкам
    folder_stats = {}
    
    # Обходим все папки и файлы
    for root, dirs, files in os.walk(base_folder):
        # Пропускаем системные папки
        dirs[:] = [d for d in dirs if not d.startswith('.') and d != 'размеченные_данные']
        
        # Ищем Excel файлы
        excel_files = [f for f in files if f.endswith('.xlsx') and not f.endswith('_разметка.xlsx')]
        
        if excel_files:
            # Создаем соответствующую структуру папок в выходной директории
            rel_path = os.path.relpath(root, base_folder)
            if rel_path == '.':
                output_subfolder = output_base_folder
            else:
                output_subfolder = os.path.join(output_base_folder, rel_path)
            
            os.makedirs(output_subfolder, exist_ok=True)
            
            print(f"\n{'='*60}")
            print(f"Обработка папки: {rel_path if rel_path != '.' else 'корневая'}")
            print(f"{'='*60}")
            
            # Статистика по текущей папке
            folder_stat = {}
            
            # Обрабатываем каждый файл двумя словарями
            for excel_file in excel_files:
                file_path = os.path.join(root, excel_file)
                file_name_without_ext = os.path.splitext(excel_file)[0]
                
                print(f"\nФайл: {excel_file}")
                
                # Обработка словарем 1 (kartaslovsent)
                print("  - Обработка словарем kartaslovsent...")
                stats1 = process_single_file(file_path, output_subfolder, 1, "kartaslovsent")
                
                # Обработка словарем 2 (rusentilex)
                print("  - Обработка словарем rusentilex...")
                stats2 = process_single_file(file_path, output_subfolder, 2, "rusentilex")
                
                # Сохраняем статистику для текущего файла
                if stats1 and stats2:
                    folder_stat[file_name_without_ext] = {
                        'kartaslovsent': stats1,
                        'rusentilex': stats2
                    }
            
            # Сохраняем статистику по папке
            if folder_stat:
                folder_stats[rel_path] = folder_stat
                save_folder_statistics(folder_stat, output_subfolder, rel_path)
    
    # Сохраняем общую статистику
    if folder_stats:
        save_overall_statistics(folder_stats, output_base_folder)
    
    print(f"\n{'='*60}")
    print("Обработка завершена!")
    print(f"Результаты сохранены в: {output_base_folder}")
    print(f"Логи сохранены в: {log_file}")
    print(f"{'='*60}")

# ===============================
# Сохранение статистики по папке
# ===============================
def save_folder_statistics(folder_stat, output_folder, folder_name):
    """Сохраняет статистику по всем файлам в папке"""
    
    # Создаем DataFrame для статистики
    data = []
    
    for file_name, stats in folder_stat.items():
        for dict_name, (total, positive, negative) in stats.items():
            data.append({
                'Файл': file_name,
                'Словарь': dict_name,
                'Всего текстов': total,
                'Позитивных': positive,
                'Негативных': negative,
                '% позитивных': positive / total * 100 if total > 0 else 0,
                '% негативных': negative / total * 100 if total > 0 else 0
            })
    
    df_stats = pd.DataFrame(data)
    
    # Сохраняем в Excel
    stats_file = os.path.join(output_folder, f"статистика_папки.xlsx")
    df_stats.to_excel(stats_file, index=False)
    
    # Создаем сводную таблицу по словарям
    pivot_file = os.path.join(output_folder, f"сводная_по_словарям.xlsx")
    pivot_df = df_stats.pivot_table(
        index='Файл',
        columns='Словарь',
        values=['Всего текстов', 'Позитивных', 'Негативных', '% позитивных', '% негативных'],
        aggfunc='first'
    )
    pivot_df.to_excel(pivot_file)
    
    print(f"  Статистика папки сохранена в: {stats_file}")
    print(f"  Сводная таблица сохранена в: {pivot_file}")

# ===============================
# Сохранение общей статистики
# ===============================
def save_overall_statistics(folder_stats, output_base_folder):
    """Сохраняет общую статистику по всем папкам"""
    
    all_data = []
    
    for folder_name, files in folder_stats.items():
        for file_name, stats in files.items():
            for dict_name, (total, positive, negative) in stats.items():
                all_data.append({
                    'Папка': folder_name if folder_name != '.' else 'корневая',
                    'Файл': file_name,
                    'Словарь': dict_name,
                    'Всего текстов': total,
                    'Позитивных': positive,
                    'Негативных': negative,
                    '% позитивных': positive / total * 100 if total > 0 else 0,
                    '% негативных': negative / total * 100 if total > 0 else 0
                })
    
    df_all = pd.DataFrame(all_data)
    
    # Сохраняем общую статистику
    overall_file = os.path.join(output_base_folder, "общая_статистика.xlsx")
    
    with pd.ExcelWriter(overall_file, engine='openpyxl') as writer:
        # Лист с детальной статистикой
        df_all.to_excel(writer, sheet_name='Детальная', index=False)
        
        # Лист со сводной по папкам
        pivot_folder = df_all.pivot_table(
            index='Папка',
            columns='Словарь',
            values=['Всего текстов', 'Позитивных', 'Негативных'],
            aggfunc='sum'
        )
        pivot_folder.to_excel(writer, sheet_name='Сводная_по_папкам')
        
        # Лист со средними процентами по папкам
        avg_percent = df_all.groupby(['Папка', 'Словарь'])[['% позитивных', '% негативных']].mean()
        avg_percent.to_excel(writer, sheet_name='Средние_проценты')
    
    print(f"\nОбщая статистика сохранена в: {overall_file}")

# ===============================
# Основная функция
# ===============================
def main():
    """
    Основная функция для запуска обработки
    """
    # Укажите путь к вашей иерархии папок
    # Например: "results_sdg" или "Астраханская область_посты_2020_filtered_20260219_011341"
    base_folder = "results_filtered/results_sdg"  # <-- ИЗМЕНИТЕ НА ВАШ ПУТЬ
    
    # Проверяем существование папки
    if not os.path.exists(base_folder):
        print(f"Ошибка: Папка {base_folder} не найдена!")
        print("Пожалуйста, укажите правильный путь к папке с данными.")
        return
    
    # Запускаем обработку
    process_all_files_in_hierarchy(base_folder)
    
    print("\nГотово! Проверьте папку 'размеченные_данные' для просмотра результатов.")

if __name__ == "__main__":
    main()