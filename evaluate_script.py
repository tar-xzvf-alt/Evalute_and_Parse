#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Batch script for processing all Excel files in a folder:
- Predict SDG labels for Russian texts
- Filter by confidence threshold
- Save simplified results (only text, predicted class, confidence)
- Generate distribution charts for both filtered and unfiltered data
"""

import pandas as pd
import numpy as np
import tensorflow as tf
import tensorflow_addons as tfa
from transformers import AutoTokenizer, TFAutoModelForSequenceClassification
from tensorflow.keras import layers
import argparse
import os
from datetime import datetime
from collections import Counter
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import glob
import warnings
warnings.filterwarnings('ignore')


def create_model(num_classes=17):
    """Create the same model architecture that was used for training"""
    
    base_model = TFAutoModelForSequenceClassification.from_pretrained(
        "DeepPavlov/rubert-base-cased",
        num_labels=num_classes,
        from_pt=True
    )
    
    input_ids = layers.Input(shape=(None,), dtype=tf.int32, name="input_ids")
    attention_mask = layers.Input(shape=(None,), dtype=tf.int32, name="attention_mask")
    
    bert_outputs = base_model(input_ids=input_ids, attention_mask=attention_mask).logits
    
    x = layers.Dense(256, activation="relu")(bert_outputs)
    x = layers.Dropout(0.3)(x)
    outputs = layers.Dense(num_classes, activation='softmax')(x)
    
    model = tf.keras.Model(inputs=[input_ids, attention_mask], outputs=outputs)
    
    return model


def load_model_with_custom_objects(model_path, num_classes):
    """Try different methods to load the model"""
    
    # Метод 1: Попробовать загрузить с custom_objects
    try:
        print("  Попытка 1: Загрузка с custom_objects...")
        model = tf.keras.models.load_model(
            model_path,
            custom_objects={
                'F1Score': tfa.metrics.F1Score,
                'Precision': tf.keras.metrics.Precision,
                'Recall': tf.keras.metrics.Recall,
                'CategoricalAccuracy': tf.keras.metrics.CategoricalAccuracy
            }
        )
        print("  ✓ Модель загружена через custom_objects")
        return model
    except Exception as e:
        print(f"  ✗ Ошибка: {e}")
    
    # Метод 2: Создать модель и загрузить веса
    try:
        print("  Попытка 2: Создание модели и загрузка весов...")
        model = create_model(num_classes)
        
        # Ищем файлы весов
        possible_weights_paths = [
            os.path.join(model_path, 'variables', 'variables'),
            os.path.join(model_path, 'variables', 'variables.data-00000-of-00001'),
            model_path
        ]
        
        weights_loaded = False
        for weights_path in possible_weights_paths:
            if os.path.exists(weights_path) or os.path.exists(weights_path + '.index'):
                model.load_weights(weights_path)
                print(f"  ✓ Веса загружены из {weights_path}")
                weights_loaded = True
                break
        
        if not weights_loaded:
            print("  ✗ Файлы весов не найдены")
            return None
        
        # Компилируем модель
        model.compile(
            optimizer=tf.keras.optimizers.Adam(learning_rate=2e-5),
            loss=tf.keras.losses.CategoricalCrossentropy(),
            metrics=[
                tf.keras.metrics.CategoricalAccuracy(name='accuracy'),
                tf.keras.metrics.Precision(name='precision', top_k=1),
                tf.keras.metrics.Recall(name='recall', top_k=1),
                tfa.metrics.F1Score(num_classes=num_classes, average='macro', name='f1')
            ]
        )
        
        return model
    except Exception as e:
        print(f"  ✗ Ошибка: {e}")
    
    return None


def load_excel_file(file_path, text_column, sheet_name=0):
    """
    Load data from Excel file
    
    Args:
        file_path: path to Excel file
        text_column: name of column with text
        sheet_name: sheet name or index
    
    Returns:
        DataFrame with loaded data or None if error
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # Проверка наличия колонки с текстом
        if text_column not in df.columns:
            print(f"    ⚠️ Колонка '{text_column}' не найдена в файле {os.path.basename(file_path)}")
            print(f"       Доступные колонки: {list(df.columns)}")
            return None
        
        # Удаляем пустые строки
        df = df.dropna(subset=[text_column])
        df = df[df[text_column].astype(str).str.strip() != '']
        
        return df
    except Exception as e:
        print(f"    ❌ Ошибка при загрузке {os.path.basename(file_path)}: {e}")
        return None


def predict_texts(model, tokenizer, texts, batch_size=24, max_length=256):
    """
    Predict labels for a list of texts
    
    Args:
        model: loaded model
        tokenizer: loaded tokenizer
        texts: list of texts to predict
        batch_size: batch size for prediction
        max_length: maximum sequence length
    
    Returns:
        predictions: array of prediction probabilities
        predicted_classes: array of predicted class indices (0-based)
        confidences: array of confidence scores
    """
    
    # Токенизация текстов
    encodings = tokenizer(
        texts,
        truncation=True,
        padding='max_length',
        max_length=max_length,
        return_tensors='tf'
    )
    
    # Подготовка входных данных
    model_inputs = {
        'input_ids': encodings['input_ids'],
        'attention_mask': encodings['attention_mask'],
    }
    
    # Добавляем token_type_ids если нужно
    if 'token_type_ids' not in encodings:
        model_inputs['token_type_ids'] = tf.zeros_like(encodings['input_ids'])
    else:
        model_inputs['token_type_ids'] = encodings['token_type_ids']
    
    # Получение предсказаний
    predictions = model.predict(model_inputs, batch_size=batch_size, verbose=0)
    
    # Получаем предсказанные классы и уверенность
    predicted_classes = np.argmax(predictions, axis=1)
    confidences = np.max(predictions, axis=1)
    
    return predictions, predicted_classes, confidences


def plot_class_distribution(class_counts, title, output_file, total_count=None):
    """
    Plot class distribution bar chart and pie chart
    
    Args:
        class_counts: Counter object with class counts
        title: chart title
        output_file: output file path
        total_count: total number of samples (for percentages)
    """
    if not class_counts:
        return
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    # Prepare data
    classes = sorted(class_counts.keys())
    counts = [class_counts[c] for c in classes]
    
    # Extract numeric values for sorting (assuming format "SDG X")
    class_numbers = []
    for c in classes:
        try:
            num = int(c.replace('SDG', '').strip())
        except:
            num = 0
        class_numbers.append(num)
    
    # Sort by class number
    sorted_data = sorted(zip(class_numbers, classes, counts))
    class_numbers, classes, counts = zip(*sorted_data) if sorted_data else ([], [], [])
    
    total = total_count if total_count else sum(counts)
    percentages = [(count / total) * 100 for count in counts]
    
    # Bar chart
    bars = ax1.bar(classes, counts, color='skyblue', edgecolor='navy', alpha=0.7)
    ax1.set_xlabel('Класс', fontsize=12)
    ax1.set_ylabel('Количество', fontsize=12)
    ax1.set_title(f'{title}\n(Всего: {total})', fontsize=14)
    ax1.tick_params(axis='x', rotation=45)
    ax1.grid(True, alpha=0.3, axis='y')
    
    # Add values on bars
    for bar, count, pct in zip(bars, counts, percentages):
        if count > 0:
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                    f'{count}\n({pct:.1f}%)', ha='center', va='bottom', fontsize=9)
    
    # Pie chart for classes with >0
    positive_data = [(c, p) for c, p, cnt in zip(classes, percentages, counts) if cnt > 0]
    if positive_data:
        pos_classes, pos_percentages = zip(*positive_data)
        
        # Use a color palette
        colors = plt.cm.Set3(np.linspace(0, 1, len(pos_classes)))
        
        wedges, texts, autotexts = ax2.pie(pos_percentages, 
                                           labels=pos_classes,
                                           autopct='%1.1f%%',
                                           startangle=90,
                                           colors=colors)
        ax2.set_title('Распределение (%)', fontsize=14)
        
        # Improve text readability
        for text in texts:
            text.set_fontsize(10)
        for autotext in autotexts:
            autotext.set_fontsize(9)
            autotext.set_color('white')
    
    plt.tight_layout()
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"    ✅ Диаграмма сохранена: {os.path.basename(output_file)}")


def process_file(file_path, model, tokenizer, text_column, threshold, 
                 num_classes, batch_size, max_length, output_dir):
    """
    Process a single Excel file
    
    Args:
        file_path: path to Excel file
        model: loaded model
        tokenizer: loaded tokenizer
        text_column: name of column with text
        threshold: confidence threshold
        num_classes: number of classes
        batch_size: batch size for prediction
        max_length: maximum sequence length
        output_dir: output directory for results
    
    Returns:
        dict with processing statistics
    """
    
    file_name = os.path.basename(file_path)
    print(f"\n  📄 Обработка файла: {file_name}")
    
    # Загрузка данных
    df = load_excel_file(file_path, text_column)
    if df is None or len(df) == 0:
        print(f"    ⚠️ Нет данных для обработки")
        return None
    
    texts = df[text_column].astype(str).tolist()
    print(f"    Загружено текстов: {len(texts)}")
    
    # Предсказание
    predictions, predicted_classes, confidences = predict_texts(
        model, tokenizer, texts, batch_size, max_length
    )
    
    # Создаем упрощенный DataFrame с результатами
    results_df = pd.DataFrame({
        'text': texts,
        'predicted_class': [f"SDG {c+1}" for c in predicted_classes],
        'confidence': confidences
    })
    
    # Фильтрация по порогу
    mask = confidences >= threshold
    filtered_df = results_df[mask].copy()
    
    # Подсчет распределения классов
    all_classes = [f"SDG {c+1}" for c in predicted_classes]
    all_class_counts = Counter(all_classes)
    
    filtered_class_counts = Counter(filtered_df['predicted_class']) if len(filtered_df) > 0 else Counter()
    
    # Статистика по файлу
    stats = {
        'file_name': file_name,
        'total_texts': len(texts),
        'filtered_texts': len(filtered_df),
        'filtered_percentage': (len(filtered_df) / len(texts)) * 100 if len(texts) > 0 else 0,
        'mean_confidence': np.mean(confidences),
        'all_class_counts': all_class_counts,
        'filtered_class_counts': filtered_class_counts
    }
    
    # Сохранение результатов
    base_name = os.path.splitext(file_name)[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Сохраняем все результаты
    all_output = os.path.join(output_dir, f"{base_name}_all_{timestamp}.xlsx")
    with pd.ExcelWriter(all_output, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='All_Predictions', index=False)
        
        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets['All_Predictions']
        
        # Заголовки
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Автоширина
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    print(f"    ✅ Сохранены все предсказания: {os.path.basename(all_output)}")
    
    # Диаграмма для всех данных
    chart_all = os.path.join(output_dir, f"{base_name}_chart_all_{timestamp}.png")
    plot_class_distribution(
        all_class_counts, 
        f"Распределение классов - все данные ({file_name})", 
        chart_all,
        len(texts)
    )
    
    # Сохраняем отфильтрованные результаты (если есть)
    if len(filtered_df) > 0:
        filtered_output = os.path.join(output_dir, f"{base_name}_filtered_{timestamp}.xlsx")
        with pd.ExcelWriter(filtered_output, engine='openpyxl') as writer:
            filtered_df.to_excel(writer, sheet_name='Filtered_Predictions', index=False)
            
            # Форматирование
            workbook = writer.book
            worksheet = writer.sheets['Filtered_Predictions']
            
            # Заголовки
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Зеленый фон для отфильтрованных данных
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            for row in range(2, len(filtered_df) + 2):
                for col in range(1, 4):
                    worksheet.cell(row=row, column=col).fill = green_fill
            
            # Автоширина
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        print(f"    ✅ Сохранены отфильтрованные: {os.path.basename(filtered_output)}")
        print(f"       Отобрано {len(filtered_df)} из {len(texts)} текстов "
              f"({stats['filtered_percentage']:.1f}%)")
        
        # Диаграмма для отфильтрованных данных
        chart_filtered = os.path.join(output_dir, f"{base_name}_chart_filtered_{timestamp}.png")
        plot_class_distribution(
            filtered_class_counts, 
            f"Распределение классов - отфильтрованные (порог {threshold:.0%})", 
            chart_filtered,
            len(filtered_df)
        )
    
    return stats


def plot_aggregated_distributions(all_stats, threshold, output_dir, timestamp):
    """
    Plot aggregated distributions across all files
    
    Args:
        all_stats: list of statistics for all processed files
        threshold: confidence threshold
        output_dir: output directory
        timestamp: timestamp for file naming
    """
    
    # Aggregate counts
    total_all_counts = Counter()
    total_filtered_counts = Counter()
    total_texts_all = 0
    total_texts_filtered = 0
    
    for stats in all_stats:
        if stats is None:
            continue
        total_all_counts.update(stats['all_class_counts'])
        total_filtered_counts.update(stats['filtered_class_counts'])
        total_texts_all += stats['total_texts']
        total_texts_filtered += stats['filtered_texts']
    
    # Plot aggregated charts
    if total_all_counts:
        chart_all_agg = os.path.join(output_dir, f"aggregated_chart_all_{timestamp}.png")
        plot_class_distribution(
            total_all_counts,
            f"Сводное распределение - ВСЕ данные ({len(all_stats)} файлов)",
            chart_all_agg,
            total_texts_all
        )
    
    if total_filtered_counts:
        chart_filtered_agg = os.path.join(output_dir, f"aggregated_chart_filtered_{timestamp}.png")
        plot_class_distribution(
            total_filtered_counts,
            f"Сводное распределение - ОТФИЛЬТРОВАННЫЕ (порог {threshold:.0%})",
            chart_filtered_agg,
            total_texts_filtered
        )


def create_summary_report(all_stats, threshold, output_dir, timestamp):
    """Create a summary report of all processed files"""
    
    summary_file = os.path.join(output_dir, f"summary_report_{timestamp}.txt")
    
    # Calculate totals
    total_texts_all = sum(s['total_texts'] for s in all_stats if s)
    total_texts_filtered = sum(s['filtered_texts'] for s in all_stats if s)
    
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("ОТЧЕТ ПО ОБРАБОТКЕ ВСЕХ ФАЙЛОВ\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"Дата обработки: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Порог уверенности: {threshold:.0%}\n")
        f.write(f"Обработано файлов: {len(all_stats)}\n\n")
        
        f.write("Детали по файлам:\n")
        f.write("-" * 80 + "\n")
        
        for stats in all_stats:
            if stats is None:
                continue
            
            f.write(f"\n📄 {stats['file_name']}:\n")
            f.write(f"  Всего текстов: {stats['total_texts']}\n")
            f.write(f"  Отфильтровано: {stats['filtered_texts']} "
                   f"({stats['filtered_percentage']:.1f}%)\n")
            f.write(f"  Средняя уверенность: {stats['mean_confidence']:.4f}\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("ИТОГО:\n")
        f.write(f"  Всего обработано текстов: {total_texts_all}\n")
        f.write(f"  Всего отфильтровано: {total_texts_filtered} "
               f"({(total_texts_filtered/total_texts_all*100):.1f}%)\n")
        f.write("=" * 80 + "\n")
    
    print(f"\n📊 Сводный отчет сохранен: {os.path.basename(summary_file)}")
    return summary_file


def main():
    print("\n" + "=" * 80)
    print("ПАКЕТНАЯ ОБРАБОТКА EXCEL ФАЙЛОВ С ПОСТРОЕНИЕМ ДИАГРАММ")
    print("=" * 80)
    
    parser = argparse.ArgumentParser(description='Batch process all Excel files in a folder')
    parser.add_argument('--input_folder', type=str, required=True,
                        help='Folder containing Excel files to process')
    parser.add_argument('--text_column', type=str, default='text',
                        help='Name of column with Russian text (default: text)')
    parser.add_argument('--model_path', type=str, 
                        default='rubert_bert_RU_35data_6_epochs',
                        help='Path to trained model')
    parser.add_argument('--model_name', type=str, 
                        default='DeepPavlov/rubert-base-cased',
                        help='HuggingFace model name')
    parser.add_argument('--num_classes', type=int, default=17,
                        help='Number of classes (default: 17)')
    parser.add_argument('--batch_size', type=int, default=24,
                        help='Batch size for prediction (default: 24)')
    parser.add_argument('--max_length', type=int, default=256,
                        help='Maximum sequence length (default: 256)')
    parser.add_argument('--threshold', type=float, default=0.9,
                        help='Confidence threshold for filtering (default: 0.9)')
    parser.add_argument('--output_folder', type=str, default=None,
                        help='Output folder for results (default: input_folder/results_TIMESTAMP)')
    
    args = parser.parse_args()
    
    # Проверка входной папки
    if not os.path.exists(args.input_folder):
        print(f"❌ ОШИБКА: Папка '{args.input_folder}' не найдена!")
        return
    
    # Проверка модели
    if not os.path.exists(args.model_path):
        print(f"❌ ОШИБКА: Папка модели '{args.model_path}' не найдена!")
        return
    
    # Поиск всех Excel файлов
    excel_files = []
    for ext in ['*.xlsx', '*.xls']:
        excel_files.extend(glob.glob(os.path.join(args.input_folder, ext)))
    
    if not excel_files:
        print(f"❌ В папке '{args.input_folder}' не найдено Excel файлов!")
        return
    
    # Создание выходной папки
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output_folder is None:
        output_folder = os.path.join(args.input_folder, f"results_{timestamp}")
    else:
        output_folder = args.output_folder
    
    os.makedirs(output_folder, exist_ok=True)
    
    print("\n📊 ПАРАМЕТРЫ ЗАПУСКА:")
    print(f"  Входная папка: {args.input_folder}")
    print(f"  Найдено Excel файлов: {len(excel_files)}")
    print(f"  Колонка с текстом: '{args.text_column}'")
    print(f"  Порог уверенности: {args.threshold:.0%}")
    print(f"  Выходная папка: {output_folder}")
    
    # Загрузка модели (один раз для всех файлов)
    print(f"\n🤖 Загрузка модели...")
    tokenizer = AutoTokenizer.from_pretrained(args.model_name)
    model = load_model_with_custom_objects(args.model_path, args.num_classes)
    
    if model is None:
        print("❌ НЕ УДАЛОСЬ ЗАГРУЗИТЬ МОДЕЛЬ!")
        return
    
    print(f"\n✅ Модель успешно загружена")
    
    # Обработка каждого файла
    all_stats = []
    
    print(f"\n📂 НАЧАЛО ОБРАБОТКИ {len(excel_files)} ФАЙЛОВ")
    print("=" * 80)
    
    for i, file_path in enumerate(excel_files, 1):
        print(f"\n[{i}/{len(excel_files)}]", end="")
        
        stats = process_file(
            file_path=file_path,
            model=model,
            tokenizer=tokenizer,
            text_column=args.text_column,
            threshold=args.threshold,
            num_classes=args.num_classes,
            batch_size=args.batch_size,
            max_length=args.max_length,
            output_dir=output_folder
        )
        
        if stats:
            all_stats.append(stats)
    
    # Построение сводных диаграмм
    if all_stats:
        print(f"\n\n📊 Построение сводных диаграмм...")
        plot_aggregated_distributions(all_stats, args.threshold, output_folder, timestamp)
    
    # Создание сводного отчета
    if all_stats:
        create_summary_report(all_stats, args.threshold, output_folder, timestamp)
    
    print("\n" + "=" * 80)
    print(f"✅ ОБРАБОТКА ЗАВЕРШЕНА!")
    print("=" * 80)
    print(f"\n📁 Результаты сохранены в папке: {output_folder}")
    print(f"   Обработано файлов: {len(all_stats)}")
    print(f"   Всего Excel файлов в папке: {len(excel_files)}")
    print(f"\n📈 Созданные диаграммы:")
    print(f"   - Для каждого файла: распределение по классам (все данные и отфильтрованные)")
    print(f"   - Сводные диаграммы: общее распределение по всем файлам")


if __name__ == "__main__":
    main()