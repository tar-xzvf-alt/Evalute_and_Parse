# coding: utf-8

import pandas as pd
import os
import sys
import logging
from openpyxl import Workbook
import pymorphy3
import matplotlib.pyplot as plt
import nltk
from nltk.corpus import stopwords

nltk.download('stopwords')

# ===============================
# Универсальный путь к ресурсам
# ===============================
def resource_path(relative_path):
    """ Получает абсолютный путь к ресурсу (работает и из .py, и из .exe) """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# ===============================
# Настройка логов
# ===============================
logging.basicConfig(
    filename='sentiment_analysis.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

morph = pymorphy3.MorphAnalyzer()
russian_stopwords = set(stopwords.words("russian"))
df_left = []

# Можно указать относительный путь к папке данных:
base_folder = resource_path("data")


# ===============================
# Загрузка словаря тональностей
# ===============================
def load_sentiment_dict(csv_path,dict_id):
    try:
        if dict_id == 1:
            sentiment_data = pd.read_csv(csv_path, usecols=['term', 'value'], sep=";")
            return dict(zip(sentiment_data["term"], sentiment_data["value"]))
        else:
            sentiment_data_1 = pd.read_csv(csv_path, usecols=['term', 'value'], sep=",")
            return dict(zip(sentiment_data_1["term"], sentiment_data_1["value"]))
    except Exception as e:
        logging.error(f"Ошибка загрузки словаря: {e}")
        raise


# ===============================
# Определение тональности текста
# ===============================
def determine_sentiment_score(text, dict_id):

    if dict_id == 1:
        csv_name = "data/kartaslovsent.csv"
    else:
        csv_name = "data/rusentilex.csv"
    sentiment_csv_path = resource_path(csv_name)
    sentiment_dict = load_sentiment_dict(sentiment_csv_path,dict_id)

    if not isinstance(text, str) or not text.strip():
        return 0.0

    words = text.split()
    score = 0
    count = 0

    for word in words:
        try:
            normal_form = morph.parse(word.lower())[0].normal_form
            if normal_form in sentiment_dict and normal_form not in russian_stopwords:
                score += sentiment_dict[normal_form]
                count += 1
        except:
            continue

    if count == 0:
        df_left.append(text)
        return 0.0

    average_score = score / count
    return 1.0 if average_score >= 0.2 else 0.0

# ===============================
# Визуализация
# ===============================
def visualize_sentiment(df, file_name):
    sentiment_counts = df["sentiment"].value_counts()
    labels = ['Негатив', 'Позитив']
    sizes = [
        sentiment_counts.get(0.0, 0),
        sentiment_counts.get(1.0, 0)
    ]

    plt.figure(figsize=(8, 6))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, colors=['#ff9999', '#99ff99'])
    plt.title(f'Распределение тональности ({file_name})')
    plt.axis('equal')

    plot_path = os.path.join(os.path.dirname(file_name), f"{os.path.splitext(os.path.basename(file_name))[0]}_plot.png")
    plt.savefig(plot_path)
    plt.close()
    logging.info(f"График сохранен: {plot_path}")

    return plot_path


def visualize_multiple(dfs, titles, base_file_path):
    # Увеличиваем размер фигуры для крупных шрифтов
    fig, axes = plt.subplots(1, len(dfs), figsize=(10 * len(dfs), 10))

    if len(dfs) == 1:
        axes = [axes]

    # Полупрозрачные цвета (alpha=0.7) в сине-зеленой гамме
    colors = ['#FF6B6BCC', '#20C997CC']  # CC в конце = 80% непрозрачности
    # Альтернативные цвета: ['#E64980CC', '#15AABFCC']

    # Параметры шрифтов
    label_font = {'size': 18, 'weight': 'bold'}
    percent_font = {'size': 16, 'weight': 'bold'}
    title_font = {'size': 20, 'weight': 'bold'}

    for ax, df, title in zip(axes, dfs, titles):
        sentiment_counts = df["sentiment"].value_counts()
        sizes = [sentiment_counts.get(0.0, 0), sentiment_counts.get(1.0, 0)]
        labels = ['Негатив', 'Позитив']

        # Создаем диаграмму с прозрачностью и толстыми разделителями
        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%',
            startangle=90,
            colors=colors,
            textprops=label_font,
            wedgeprops={'linewidth': 2, 'edgecolor': 'white', 'alpha': 0.8},
            pctdistance=0.8
        )

        # Устанавливаем стиль для процентов внутри
        for autotext in autotexts:
            autotext.set_fontsize(percent_font['size'])
            autotext.set_fontweight(percent_font['weight'])
            autotext.set_color('white')  # Белый текст для лучшей читаемости

        # Настраиваем заголовок
        ax.set_title(title, fontdict=title_font, pad=25)


    plt.tight_layout(pad=3.0)

    plot_path = os.path.join(os.path.dirname(base_file_path),
                             f"{os.path.splitext(os.path.basename(base_file_path))[0]}_multi_plot.png")
    plt.savefig(plot_path, dpi=150, bbox_inches='tight', transparent=True)
    plt.close()
    return plot_path

# ===============================
# Анализ всех файлов в папке
# ===============================
def analyse_combined(base_folder_analyse):

    for file_name in os.listdir(base_folder_analyse):
        folder_path = base_folder_analyse + '/' + file_name

        for file_name in os.listdir(base_folder_analyse):
            #if not file_name.endswith('посты.xlsx') or not file_name.endswith('комментарии.xlsx'):
            #    continue

            print(f"Обрабатываем файл: {file_name}")
            logging.info(f"Обработка файла: {file_name}")
            file_path = os.path.join(folder_path, file_name)

            try:
                df = pd.read_excel(file_path)
                if df.empty:
                    logging.warning(f"Файл пуст: {file_name}")
                    continue

                text_column = df.columns[0]
                df = df.rename(columns={text_column: 'text'})
                df = df.dropna(subset=['text'])
                df['text'] = df['text'].astype(str)
                df = df[df['text'].str.strip().astype(bool)]

                df["sentiment"] = df["text"].apply(lambda t: determine_sentiment_score(t, dict_id=1))

                if df_left:
                    df = df[~df["text"].isin(df_left)]

                output_name = os.path.splitext(file_name)[0]
                output_path = os.path.join(folder_path, f"{output_name}_разметка.xlsx")
                df.to_excel(output_path, index=False)
                logging.info(f"Результаты сохранены: {output_path}")

                visualize_sentiment(df, file_path)

                total = len(df)
                positive = (df["sentiment"] == 1.0).sum()
                negative = (df["sentiment"] == 0.0).sum()

                wb = Workbook()
                ws = wb.active
                stats = [
                    ("Файл", output_name),
                    ("Всего текстов", total),
                    ("Негативных", negative),
                    ("Позитивных", positive),
                    ("% негативных", negative / total * 100),
                    ("% позитивных", positive / total * 100)
                ]

                for row, (label, value) in enumerate(stats, start=1):
                    ws[f"A{row}"] = label
                    ws[f"B{row}"] = value

                stats_path = os.path.join(folder_path, f"{output_name}_итоги.xlsx")
                wb.save(stats_path)
                logging.info(f"Статистика сохранена: {stats_path}")

            except Exception as e:
                logging.error(f"Ошибка обработки файла {file_name}: {e}")
                continue

    logging.info("Обработка завершена!")
    print("Анализ завершен. Проверьте папки с результатами.")


#analyse_combined("results_sdg\Астраханская область_посты_2020_filtered_20260219_011341")